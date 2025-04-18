import pandas as pd
from fuzzywuzzy import fuzz, process
import difflib
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def common_prefix(strings):
    if not strings:
        return ""
    min_str = min(strings, key=len)
    for i in range(len(min_str)):
        for s in strings:
            if s[i] != min_str[i]:
                return min_str[:i]
    return min_str

# Function to clean descriptions by removing common prefixes
def clean_descriptions(terms):
    grouped_by_word_count = defaultdict(list)
    
    # Group descriptions by word count, ensuring valid strings
    for term in terms:
        if isinstance(term, str):  # Only process valid strings
            word_count = len(term.split())
            grouped_by_word_count[word_count].append(term)
    
    # Remove common prefix for each group
    cleaned_terms = {}
    for word_count, group in grouped_by_word_count.items():
        if len(group) > 1:  # Only process groups with more than one term
            try:
                common_prefix_str = common_prefix(group)
                cleaned_terms[word_count] = [term[len(common_prefix_str):].strip() for term in group]
            except Exception as e:
                print(f"Error processing group with word count {word_count}: {e}")
        else:
            # Handle single-term groups by just copying them without modification
            cleaned_terms[word_count] = group
    
    return cleaned_terms

# Function to filter the a subset of a DataFrame based on a search string
def filter_df(df, input_string):
    if 'HEADER NAME' not in df.columns:
        print("The 'HEADER NAME' column was not found.")
        return pd.DataFrame()
    
    filtered_df = df[df['HEADER NAME'].str.contains(input_string, case=False, na=False, regex=False)]

    return filtered_df

# Function to match rows between two DataFrames using fuzzy matching (partial ratio)
def find_closest_match(description, choices):
    match = process.extractOne(description, choices, scorer=fuzz.partial_ratio)
    return match

def compare(acmv_df, d3_df, prefix):
    # Add new columns to acmv_df to store the corresponding rates and descriptions
    acmv_df = acmv_df[["HEADER NAME", "DESCRIPTION", "UNIT", "RATE"]].copy()
    d3_df = d3_df.copy()
    pdesc =f"{prefix}"
    prate =f"{prefix} RATE"

    acmv_df[pdesc] = None
    acmv_df[prate] = None
    acmv_df['Score'] = None
    #acmv_df["Clean"] = None
    print(acmv_df.shape[0])
    print(d3_df.shape[0])

    cleaned_terms = clean_descriptions(d3_df['DESCRIPTION'])

    # Create iterators for each word_count group
    term_iterators = {wc: iter(terms) for wc, terms in cleaned_terms.items()}

    def get_cleaned_description(description):
        if isinstance(description, str):
            word_count = len(description.split())
            iterator = term_iterators.get(word_count)
            if iterator:
                try:
                    return next(iterator)
                except StopIteration:
                    return description
        return description


    d3_df['Clean Description'] = d3_df['DESCRIPTION'].apply(get_cleaned_description)

    # List to keep track of the matched descriptions from d3_df
    matched_descriptions = []

    # Iterate through each row of acmv.csv and find the best match from d3.csv
    for index, row in acmv_df.iterrows():
        acmv_description = str(row['DESCRIPTION'])
        if pd.isna(acmv_description) or not str(acmv_description).strip():  # NaN or empty string after stripping
            continue
        d3_descriptions = d3_df['Clean Description'].tolist()
        
        # Get the closest match
        match = find_closest_match(acmv_description, d3_descriptions)
        if match == None:
            continue
        # Get the corresponding rate from d3_df

        matching_row = d3_df[d3_df['Clean Description'] == match[0]]
        
        if not matching_row.empty:  # If there is a corresponding rate in d3_df
            rate = matching_row['RATE'].values[0]
            desc = matching_row['DESCRIPTION'].values[0]
            # Update the 'D3 Description', 'D3 Rate', and 'Score' columns in acmv_df
            acmv_df.at[index, pdesc] = desc
            acmv_df.at[index, prate] = rate
            acmv_df.at[index, 'Score'] = match[1]
            #acmv_df.at[index, "Clean"] = match[0]
            
            # Append the matched description to the list
            matched_descriptions.append(desc)

        #debug
        #print(f"ACMV Description: {acmv_description}")
        #print(f"Closest Match from d3.csv: {match[0]}")
        #print(f"Match Score: {match[1]}")
        #print(f"Corresponding d3.csv Rate: {rate if not matching_row.empty else 'N/A'}")
        #print("-" * 80)  # Separator line for readability

    # Create the copied_df by filtering d3_df for the matched descriptions
    copied_d3 = d3_df[d3_df['DESCRIPTION'].isin(matched_descriptions)].copy()
    return acmv_df, copied_d3


def check(updated_acmv_df, filtered_acmv, filtered_d3, copied_d3, prefix):
    # Initialize 'Check' column as None (empty or NaN initially)
    pdesc =f"{prefix}"
    prate =f"{prefix} RATE"
    d3_not_in_acmv = pd.DataFrame()
    
    if not updated_acmv_df.empty:
        updated_acmv_df = updated_acmv_df.copy()
        updated_acmv_df.loc[:, "Check"] = None  # Start with None (can be changed to empty string if needed)

        # Convert 'Score' to numeric, coercing errors to NaN
        updated_acmv_df.loc[:, "Score"] = pd.to_numeric(updated_acmv_df["Score"], errors="coerce")
        updated_acmv_df.loc[:, "RATE"] = pd.to_numeric(updated_acmv_df["RATE"], errors="coerce")
        updated_acmv_df.loc[:, prate] = pd.to_numeric(updated_acmv_df[prate], errors="coerce")
        
        # Append 'Score' check to 'Check' column where condition is met
        updated_acmv_df.loc[updated_acmv_df["Score"] <= 60, "Check"] = updated_acmv_df.loc[updated_acmv_df["Score"] <= 60, "Check"].apply(lambda x: f"{x}, Score" if x and isinstance(x, str) else "Score")
        
        # Append 'Price Difference' check to 'Check' column where condition is met, but not if 'labour' is in 'DESCRIPTION'
        updated_acmv_df.loc[
            (abs(updated_acmv_df["RATE"] - updated_acmv_df[prate]) / updated_acmv_df["RATE"] > 0.50) &
            (~updated_acmv_df[pdesc].str.contains("labour", case=False, na=False)), "Check"] = updated_acmv_df.loc[
            (abs(updated_acmv_df["RATE"] - updated_acmv_df[prate]) / updated_acmv_df["RATE"] > 0.50) &
            (~updated_acmv_df[pdesc].str.contains("labour", case=False, na=False)), "Check"].apply(
            lambda x: f"{x}, Price" if x and isinstance(x, str) else "Price")
        
        # Case 1: acmv > d3 (if the difference in the number of rows is positive)
        diff = filtered_acmv.shape[0] - filtered_d3.shape[0]
        if diff > 0:
            # Get the minimum value of the 'Score' column
            min_score = updated_acmv_df["Score"].min()
            # Append 'Too many ACMV' to 'Check' column where the score equals the minimum score

            updated_acmv_df.loc[updated_acmv_df["Score"] == min_score, "Check"] = updated_acmv_df.loc[
                updated_acmv_df["Score"] == min_score, "Check"].apply(lambda x: f"{x}, Too many ACMV" if x and isinstance(x, str) else "Too many ACMV")
            updated_acmv_df.loc[updated_acmv_df["Score"] == min_score, [pdesc, prate]] = None
            
        elif diff < 0:
            d3_not_in_acmv = filtered_d3[~filtered_d3.index.isin(copied_d3.index)]
    return updated_acmv_df, d3_not_in_acmv

#reorder columns to kylee's standards dynamically
def reorder(df):
    constlist = [0, 1]
    for i in range(4, df.shape[1], 8):
        constlist.append(i)
    
    constlist.append(2)
    
    for j in range(df.shape[1]//8):
        constlist.append(3 + 8 * j)
        for i in range(5, 8):
            constlist.append(i + 8 * j)

    reordered_df = df.iloc[:, constlist]
    return reordered_df
     
#in the case of an error whereby there is no Description in "HEADER COMPARIOSN",
def empty(acmv_df, prefix):
    acmv_df = acmv_df[["HEADER NAME", "DESCRIPTION", "UNIT", "RATE"]].copy()
    pdesc =f"{prefix}"
    prate =f"{prefix} RATE"

    acmv_df[pdesc] = None
    acmv_df[prate] = None
    acmv_df['Score'] = None
    #acmv_df["Clean"] = None
    return acmv_df

def main(input, output):
    #output file
    date = get_today_date()
    #initialise excel sheet
    hello = pd.DataFrame()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        hello.to_excel(writer, sheet_name='Sheet1', index=False)

    database = pd.DataFrame()

    #BASE FILE (input)
    acmv_df = pd.read_excel(input, sheet_name='INPUT 1 (ACMV)')
    #HEADER COMPARISON
    ls = pd.read_excel(input, sheet_name='HEADER COMPARISON')
    
    for i in range(ls.shape[1]-2):
        temp_acmv = []
        temp_copied = []
        if not "SOR" in ls.columns[i+2].upper():
            print("In this Header Comparison file, there are no 'SORS' ")
            continue
        #copy sheets
        d3_df = pd.read_excel(input, sheet_name=ls.columns[i+2].strip())
        for index, row in ls.iterrows():
            acmv_str = row.iloc[1]
            d3_str = row.iloc[i+2]
            print("acmv_str is", acmv_str)
            print("d3_str is", d3_str)
            prefix = ls.columns[i+2].strip()

            #end of file
            if pd.isna(d3_str) and pd.isna(acmv_str):
                break
            
            #error check for d3_line empty
            if pd.isna(d3_str):
                filtered_acmv = filter_df(acmv_df, acmv_str)
                updated_acmv_df = empty(filtered_acmv, prefix)
                temp_acmv.append(updated_acmv_df)
                continue

            elif pd.isna(acmv_str):
                print("HEADER COMPARISON ACMV VALUE MISSING")
                continue

            filtered_acmv = filter_df(acmv_df, acmv_str)
            filtered_d3 = filter_df(d3_df, d3_str)
            # Perform the matching between the filtered DataFrames
            updated_acmv_df, copied_d3 = compare(filtered_acmv, filtered_d3, prefix)
            updated_acmv_df, d3_extras = check(updated_acmv_df, filtered_acmv, filtered_d3, copied_d3, prefix)
            temp_copied.append(d3_extras)
            temp_acmv.append(updated_acmv_df)
            
        #excess files for when d3>acmv
        d3_additional = pd.concat(temp_copied, ignore_index=True)
        d3_additional.columns = d3_additional.columns.str.upper()
        headers = ["HEADER NAME", "DESCRIPTION","UNIT", "RATE"]
        d3_additional = d3_additional[headers]
        with pd.ExcelWriter(output, engine='openpyxl', mode='a') as writer:
            d3_additional.to_excel(writer, sheet_name=f"SOR {i+1} Additionals", index=False)
        final = pd.concat(temp_acmv, ignore_index=True)
        #concat db if not first
        if i == 0:
            database = final
        else: 
            database = pd.concat([database, final],axis=1)
    
    with pd.ExcelWriter(output, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        database = reorder(database)
        # Remove columns with names that start with 'Unnamed:'
        database.to_excel(writer, sheet_name="ACMV", index=False)


                
def color_check_cells(file_path="output.xlsx"):

    wb = load_workbook(file_path)

    if 'Sheet1' in wb.sheetnames:
        # Remove the sheet called 'Sheet1'
        sheet_to_remove = wb['Sheet1']
        wb.remove(sheet_to_remove)

    sheet_names = wb.sheetnames
    last_sheet = wb[sheet_names[-1]] 
    print(last_sheet) # Reordering sheets
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(last_sheet))) 

    if "ACMV" not in wb.sheetnames:
        print("ACMV sheet not found in workbook")
        return
    ws = wb["ACMV"]
    # Identify all columns with headers containing "Check" (case-insensitive)
    check_columns = []
    for cell in ws[1]:
        if cell.value and "check" in str(cell.value).lower():
            check_columns.append(cell.column)
    
    if not check_columns:
        print("No Check columns found in ACMV sheet")
        return

    # Define fill styles
    purple_fill = PatternFill("solid", fgColor="d6b6d6")  # For Score diff only
    blue_fill = PatternFill("solid", fgColor="b6c9d6")      # For Price diff only
    yellow_fill = PatternFill("solid", fgColor="f7f7be")    # For both Score and Price diffs
    remove_fill = PatternFill("solid", fgColor="92d050")              # To green fill
    
    # Iterate over rows (starting from row 2, assuming row 1 is the header)
    for row in range(2, ws.max_row + 1):
        for col in check_columns:
            check_cell = ws.cell(row=row, column=col)
            if check_cell.value is not None:
                text = str(check_cell.value).lower()
                
                # If "Too many ACMV" is present, remove fill
                if "too many acmv" in text:
                    fill = remove_fill
                else:
                    has_score = "score" in text
                    has_price = "price" in text
                    if has_score and has_price:
                        fill = yellow_fill
                    elif has_score:
                        fill = purple_fill
                    elif has_price:
                        fill = blue_fill

                # Apply the fill to the check cell and the three cells immediately to its left (if available)
                for c in range(max(1, col - 3), col + 1):
                    ws.cell(row=row, column=c).fill = fill
                        
    wb.save(file_path)



# Function to return today's date in dd/mmm/yy format
def get_today_date():
    return datetime.today().strftime('%d_%b_%y')
date = get_today_date()

def copy_sheet(input, output):
    all_sheets = pd.read_excel(input, sheet_name=None)  # Returns a dict: {sheet_name: DataFrame}
    def remove_unnamed(df):
        return df.loc[:, ~df.columns.str.contains('^Unnamed', regex=True)]
    # Try to open existing destination workbook
    with pd.ExcelWriter(output, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        for sheet_name, df in all_sheets.items():
            cleaned = remove_unnamed(df)
            cleaned.to_excel(writer, sheet_name=sheet_name, index=False)

# Run the main process and then color cells as needed
if __name__ == "__main__":
    input = 'acmv_final.xlsx'
    output = f"ACMV_{get_today_date()}.xlsx"
    main(input, output)
    color_check_cells(output)
    copy_sheet(input, output)

